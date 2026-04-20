"""
Utilidades compartidas: URLs del portal DIAN, sanitización de nombres de archivo
y parseo de archivos con CUFEs (facturas).

Heredado de E:/BATUTA PROJECTS/DIAN DOWNLOEADER/core/utils.py. Extendido para
soportar .xlsx como formato de entrada de CUFEs (el equipo contable del cliente
suele exportar desde Excel).
"""

from __future__ import annotations

import os
import re
from pathlib import Path

DIAN_BASE_URL = "https://catalogo-vpfe.dian.gov.co"
DIAN_RECEIVED_URL = f"{DIAN_BASE_URL}/Document/Received"
DIAN_EXPORT_URL = f"{DIAN_BASE_URL}/Document/Export"

# Un CUFE tiene 96 chars hex pero aceptamos 50+ para tolerar formatos parciales
# o CUDE (documentos soporte), que comparten alfabeto hex.
_CUFE_RE = re.compile(r"[a-fA-F0-9]{50,}")


def sanitize_filename(name: str, max_len: int = 150) -> str:
    """Remove illegal filesystem characters from a filename.

    Args:
        name: Nombre propuesto (puede contener : / \\ | etc.).
        max_len: Longitud máxima del resultado.

    Returns:
        Nombre seguro para Windows/Linux/macOS, truncado a ``max_len`` chars.
    """
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)[:max_len]


def parse_cufe_file(filepath: str | os.PathLike) -> list[str]:
    """Extract CUFE hex strings from a text, markdown or Excel file.

    Soporta:
    - ``.txt``, ``.md``, ``.csv``: regex sobre todo el contenido.
    - ``.xlsx``: lee la primera hoja y busca CUFEs en todas las celdas.

    Args:
        filepath: Ruta al archivo con CUFEs (uno por línea o mezclados).

    Returns:
        Lista de strings hex (50+ chars). Vacía si el archivo no existe o no hay CUFEs.
    """
    path = Path(filepath)
    if not path.exists():
        return []

    if path.suffix.lower() == ".xlsx":
        return _parse_cufe_xlsx(path)

    with path.open("r", encoding="utf-8", errors="ignore") as f:
        content = f.read()
    return _CUFE_RE.findall(content)


def _parse_cufe_xlsx(path: Path) -> list[str]:
    """Parse CUFEs from an Excel .xlsx file by scanning every cell."""
    # Import local para que el módulo funcione aun si openpyxl no está instalado
    # y el usuario no usa el modo batch Excel.
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    cufes: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                cufes.extend(_CUFE_RE.findall(str(cell)))
    return cufes


def build_output_dir(base: str | os.PathLike, start_date: str, company: str) -> str:
    """Build output directory: ``base/year/company``.

    Args:
        base: Ruta raíz de descargas (ej. ``~/Documents/DIAN``).
        start_date: Fecha en formato ``YYYY/MM/DD`` (se extrae el año).
        company: Nombre de la empresa (se usa como subdirectorio).

    Returns:
        Ruta absoluta como string.
    """
    year = start_date.split("/")[0]
    return str(Path(base) / year / company)


def default_output_base() -> str:
    """Retorna ~/Documents/DIAN como raíz de descargas por defecto.

    Razón: en Windows un cliente final no tiene por qué tener una unidad E:\\
    como asumía el proyecto viejo. ``~/Documents`` siempre existe y está indexado
    por el Explorador.
    """
    return str(Path.home() / "Documents" / "DIAN")


def appdata_dir() -> Path:
    """Carpeta de datos de la app en %APPDATA%/BatutaAI/DianDownloader.

    Crea la carpeta si no existe. Se usa para checkpoints, licencia y logs.
    """
    if os.name == "nt":
        root = Path(os.environ.get("APPDATA", Path.home() / "AppData" / "Roaming"))
    else:
        # Fallback para dev en Linux/macOS
        root = Path.home() / ".config"
    path = root / "BatutaAI" / "DianDownloader"
    path.mkdir(parents=True, exist_ok=True)
    return path
