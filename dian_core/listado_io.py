"""
Utilidades para leer/escribir el Excel del listado DIAN.

Este módulo es consumido tanto por `faltantes_report` (cruce listado vs zips)
como por `listado_updater` (log de control). Extraemos la lógica aquí para
evitar duplicación y centralizar la heurística de detección de la columna
CUFE (DIAN cambió el header varias veces: "CUFE/CUDE", "Clave Única...").
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

# CUFE = 96 chars hex. Aceptamos 50+ para tolerar CUDE y formatos parciales.
_HEX_RE = re.compile(r"[a-fA-F0-9]{50,}")


def extract_cufes_from_xlsx(path: Path | str) -> list[tuple[str, dict]]:
    """Devuelve [(cufe, row_dict), ...] del Excel DIAN.

    Args:
        path: ruta al Excel generado por ListadoExporter.

    Returns:
        Lista de tuplas. ``row_dict`` es `{header: valor}` para que el caller
        pueda mostrar info contextual (número de factura, emisor, etc.).
    """
    wb = load_workbook(Path(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    cufe_col_idx = _detect_cufe_column(headers)

    results: list[tuple[str, dict]] = []
    for raw in rows[1:]:
        if raw is None:
            continue
        row_dict = {
            headers[i]: raw[i] if i < len(raw) else None
            for i in range(len(headers))
        }

        cufe = _row_cufe(raw, cufe_col_idx)
        if cufe:
            results.append((cufe, row_dict))
    return results


def iter_cufe_rows(path: Path | str) -> Iterable[tuple[str, int]]:
    """Generador de (cufe, row_index_1based) para escritura (non read-only).

    Útil para el updater que necesita row_index para actualizar celdas.
    """
    wb = load_workbook(Path(path), read_only=False, data_only=True)
    ws = wb.active
    headers = [str(c.value) if c.value is not None else f"col_{i}" for i, c in enumerate(ws[1])]
    cufe_col_idx = _detect_cufe_column(headers)

    for row in ws.iter_rows(min_row=2):
        cells = [c.value for c in row]
        cufe = _row_cufe(cells, cufe_col_idx)
        if cufe:
            yield cufe, row[0].row


def _detect_cufe_column(headers: list[str]) -> int | None:
    """Localiza la columna que contiene CUFEs por el nombre del header.

    DIAN usa distintos nombres según el tipo de doc: 'CUFE/CUDE',
    'CUFE/CUDE/CUDS', 'Clave Única...'.
    """
    for i, h in enumerate(headers):
        low = h.lower()
        if "cufe" in low or "cude" in low or "cuds" in low or "clave" in low:
            return i
    return None


def _row_cufe(row: tuple, cufe_col_idx: int | None) -> str | None:
    """Extrae el CUFE de una fila. Prueba primero la columna detectada; si no
    hay match, escanea TODAS las celdas buscando el patrón hex."""
    if cufe_col_idx is not None and cufe_col_idx < len(row):
        val = row[cufe_col_idx]
        if val is not None:
            m = _HEX_RE.search(str(val))
            if m:
                return m.group(0)

    # Fallback: escanear todas las celdas
    for cell in row:
        if cell is None:
            continue
        m = _HEX_RE.search(str(cell))
        if m:
            return m.group(0)
    return None
