"""
Generador del reporte de facturas faltantes.

Cruza la lista oficial DIAN (Excel generado por ``ListadoExporter``) contra
los archivos .zip que quedaron en la carpeta destino, y produce:
  - ``faltantes.xlsx``: filas del listado DIAN cuyo CUFE NO aparece entre los
    nombres de archivo descargados.
  - ``reporte.xlsx``: resumen ejecutivo (total DIAN, descargados, faltantes,
    porcentajes).

Al quedar junto a los ZIPs, el usuario ve el estado del job en la misma
carpeta donde descargó — sin necesidad de entrar al log.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable, Optional

from openpyxl import Workbook, load_workbook

from dian_core.utils import sanitize_filename


_HEX_RE = re.compile(r"[a-fA-F0-9]{50,}")


@dataclass
class FaltantesReport:
    """Resumen del cruce."""

    output_dir: str
    total_dian: int
    total_downloaded: int
    total_missing: int
    faltantes_file: Optional[str]
    reporte_file: Optional[str]
    missing_cufes: list[str]

    @property
    def coverage_pct(self) -> float:
        if self.total_dian == 0:
            return 100.0
        return round(100 * self.total_downloaded / self.total_dian, 1)


# ---------- helpers ----------

def _extract_cufes_from_xlsx(path: Path) -> list[tuple[str, dict]]:
    """Extrae CUFEs del Excel DIAN y devuelve [(cufe, row_dict), ...].

    El Excel DIAN tiene headers en la fila 1 y datos desde la 2. El CUFE suele
    estar en una columna con nombre como 'CUFE/CUDE' o similar. Como fallback,
    escaneamos TODAS las celdas de la fila buscando el patrón hex.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]

    # Intentamos localizar la columna del CUFE por nombre
    cufe_col_idx: Optional[int] = None
    for i, h in enumerate(headers):
        low = h.lower()
        if "cufe" in low or "cude" in low or "clave" in low:
            cufe_col_idx = i
            break

    results: list[tuple[str, dict]] = []
    for raw in rows[1:]:
        if raw is None:
            continue
        row_dict = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}

        cufe: Optional[str] = None
        if cufe_col_idx is not None and cufe_col_idx < len(raw):
            val = raw[cufe_col_idx]
            if val is not None:
                m = _HEX_RE.search(str(val))
                if m:
                    cufe = m.group(0)

        if not cufe:
            # Fallback: escanear todas las celdas
            for cell in raw:
                if cell is None:
                    continue
                m = _HEX_RE.search(str(cell))
                if m:
                    cufe = m.group(0)
                    break

        if cufe:
            results.append((cufe, row_dict))
    return results


def _downloaded_cufes_from_dir(output_dir: Path) -> set[str]:
    """Extrae CUFEs hex de los nombres de los .zip descargados.

    Los descargadores nombran los ZIPs con el row_text del portal (que incluye
    el CUFE). Hacemos regex sobre cada filename para capturarlo.
    """
    cufes: set[str] = set()
    for p in output_dir.glob("*.zip"):
        for match in _HEX_RE.finditer(p.name):
            cufes.add(match.group(0))
    return cufes


# ---------- main ----------

def generate_report(
    output_dir: str,
    listado_xlsx: Optional[str] = None,
    callback: Optional[Callable[..., None]] = None,
) -> FaltantesReport:
    """Genera ``faltantes.xlsx`` + ``reporte.xlsx`` dentro de ``output_dir``.

    Args:
        output_dir: carpeta con los .zip descargados.
        listado_xlsx: Excel DIAN a usar como fuente de verdad. Si es ``None``,
            se busca automáticamente el archivo ``listado_*.xlsx`` más reciente
            dentro de ``output_dir``.
        callback: ``(event_type, message, **kwargs)`` opcional.

    Returns:
        FaltantesReport con conteos + rutas de los dos Excel generados.
    """
    emit = callback or (lambda e, m, **kw: None)
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    # Localizar el listado
    src: Optional[Path]
    if listado_xlsx:
        src = Path(listado_xlsx)
    else:
        src = _find_latest_listado(out)

    if not src or not src.exists():
        emit(
            "error",
            "No encontré ningún listado_*.xlsx. Descarga el Listado Excel "
            "primero para poder generar el reporte de faltantes.",
        )
        return FaltantesReport(
            output_dir=str(out),
            total_dian=0,
            total_downloaded=0,
            total_missing=0,
            faltantes_file=None,
            reporte_file=None,
            missing_cufes=[],
        )

    emit("status", f"▶ Cruzando listado DIAN vs descargados")
    emit("status", f"  Listado:    {src}")
    emit("status", f"  Carpeta:    {out}")

    dian_rows = _extract_cufes_from_xlsx(src)
    downloaded = _downloaded_cufes_from_dir(out)

    emit("status", f"  DIAN:       {len(dian_rows)} filas")
    emit("status", f"  Descargados: {len(downloaded)} archivos .zip")

    missing: list[tuple[str, dict]] = [
        (cufe, row) for cufe, row in dian_rows if cufe not in downloaded
    ]
    emit("status", f"  Faltantes:  {len(missing)}")

    # Escribir Excel de faltantes
    faltantes_path = out / "faltantes.xlsx"
    _write_faltantes_xlsx(faltantes_path, missing)

    # Escribir Excel de reporte resumen
    reporte_path = out / "reporte.xlsx"
    _write_reporte_xlsx(
        reporte_path,
        total_dian=len(dian_rows),
        total_downloaded=len(downloaded),
        total_missing=len(missing),
        listado_source=str(src),
    )

    emit(
        "success",
        f"Reporte generado: {faltantes_path.name} · {reporte_path.name}",
        faltantes_file=str(faltantes_path),
        reporte_file=str(reporte_path),
        total_missing=len(missing),
    )

    return FaltantesReport(
        output_dir=str(out),
        total_dian=len(dian_rows),
        total_downloaded=len(downloaded),
        total_missing=len(missing),
        faltantes_file=str(faltantes_path),
        reporte_file=str(reporte_path),
        missing_cufes=[c for c, _ in missing],
    )


def _find_latest_listado(folder: Path) -> Optional[Path]:
    candidates = sorted(folder.glob("listado_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def _write_faltantes_xlsx(path: Path, missing: list[tuple[str, dict]]) -> None:
    """Genera un Excel con los CUFEs que faltan por descargar."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FALTANTES"

    if missing:
        headers = list(missing[0][1].keys())
        ws.append(["CUFE"] + headers)
        for cufe, row in missing:
            ws.append([cufe] + [row.get(h) for h in headers])
    else:
        ws.append(["CUFE"])
        ws.append(["— sin faltantes —"])

    wb.save(path)


def _write_reporte_xlsx(
    path: Path,
    total_dian: int,
    total_downloaded: int,
    total_missing: int,
    listado_source: str,
) -> None:
    """Resumen ejecutivo del job en una sola hoja."""
    wb = Workbook()
    ws = wb.active
    ws.title = "REPORTE"

    pct = round(100 * total_downloaded / total_dian, 1) if total_dian else 100.0

    rows = [
        ("Reporte de descarga DIAN",        ""),
        ("Generado",                         datetime.now().isoformat(timespec="seconds")),
        ("Listado fuente",                   listado_source),
        ("",                                 ""),
        ("Total en listado DIAN",            total_dian),
        ("Archivos descargados",             total_downloaded),
        ("Faltantes",                        total_missing),
        ("Cobertura",                        f"{pct}%"),
    ]
    for k, v in rows:
        ws.append([k, v])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    wb.save(path)
