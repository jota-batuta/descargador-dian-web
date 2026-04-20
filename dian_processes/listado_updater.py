"""
Listado Updater — el Excel oficial DIAN se convierte en el log de control
del job desatendido.

Mecánica:
- Al init: agrega columna ``Estado de descarga`` al extremo derecho.
  Valor="Pendiente" para filas que aún no tengan valor.
- Durante el job: cada worker llama ``mark_downloaded(cufe, ok, error)``.
  Los cambios se acumulan en memoria. Timestamp y mensaje de error van a
  ``app.log`` vía ``logging`` — no ensucian el Excel que el cliente abre.
- Un timer background hace flush cada 10 s (o si el buffer excede 20 filas).
- ``finish()`` al terminar el job garantiza flush final.

El mismo Excel SIRVE como checkpoint para resume: al re-correr el job, si
la columna ya existe, ``get_pending_cufes()`` devuelve solo filas con
Estado ∈ {Pendiente, Error}. No hay archivo de checkpoint separado.
"""

from __future__ import annotations

import logging
import os
import threading
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from dian_core.listado_io import _detect_cufe_column, _row_cufe

log = logging.getLogger(__name__)

COL_ESTADO_DESCARGA = "Estado de descarga"

STATUS_PENDIENTE = "Pendiente"
STATUS_DESCARGADO = "Descargado"
STATUS_ERROR = "Error"

_FLUSH_EVERY_N_CHANGES = 20
_FLUSH_INTERVAL_S = 10.0


class ListadoUpdater:
    """Mantiene el Excel actualizado durante el job desatendido.

    Uso:

        updater = ListadoUpdater(xlsx_path)
        pending = updater.init()                 # agrega columna, cuenta
        # ... workers concurrentes llaman:
        updater.mark_downloaded(cufe, ok=True)
        # ...
        updater.finish()                          # flush + cierra
    """

    def __init__(self, xlsx_path: str | Path) -> None:
        self.path = str(xlsx_path)
        self._wb = None
        self._ws = None
        self._lock = threading.Lock()
        self._cufe_to_row: dict[str, int] = {}
        self._cufe_to_meta: dict[str, dict] = {}
        self._col_estado: int = 0
        self._dirty_count: int = 0
        self._timer: threading.Timer | None = None
        self._closed = False

    # ---------- init / finish ----------

    def init(self) -> list[str]:
        """Abre el wb, agrega columna si falta, indexa CUFEs + metadata.

        Returns:
            Lista de CUFEs con estado ``Pendiente`` o ``Error`` (a descargar).
            Los ``Descargado`` se excluyen (resume automático).
        """
        self._wb = load_workbook(self.path)
        self._ws = self._wb.active

        headers = [
            str(c.value) if c.value is not None else f"col_{i}"
            for i, c in enumerate(self._ws[1])
        ]

        self._col_estado = self._ensure_column(headers, COL_ESTADO_DESCARGA)

        cufe_col_idx = _detect_cufe_column(headers)

        pending: list[str] = []
        for row in self._ws.iter_rows(min_row=2):
            cells = [c.value for c in row]
            cufe = _row_cufe(tuple(cells), cufe_col_idx)
            if not cufe:
                continue

            row_idx = row[0].row
            self._cufe_to_row[cufe] = row_idx

            # Snapshot the full row as {header: value} so callers (e.g. the
            # download step) can build descriptive filenames from Prefijo /
            # Folio / NIT Emisor / Nombre Emisor without re-reading the Excel.
            self._cufe_to_meta[cufe] = {
                headers[i]: cells[i] if i < len(cells) else None
                for i in range(len(headers))
                if headers[i] != COL_ESTADO_DESCARGA
            }

            estado_cell = self._ws.cell(row=row_idx, column=self._col_estado)
            estado = (estado_cell.value or "").strip() if estado_cell.value else ""

            if estado == STATUS_DESCARGADO:
                continue

            if not estado:
                estado_cell.value = STATUS_PENDIENTE
                self._dirty_count += 1
            pending.append(cufe)

        self._save()
        self._schedule_timer()
        return pending

    def get_meta(self, cufe: str) -> dict:
        """Return the row metadata captured at init for `cufe`, or ``{}``."""
        return self._cufe_to_meta.get(cufe, {})

    def finish(self) -> None:
        """Cancela timer, hace flush final, cierra workbook."""
        with self._lock:
            if self._closed:
                return
            self._closed = True

        if self._timer:
            self._timer.cancel()

        try:
            self._save()
        finally:
            if self._wb:
                try:
                    self._wb.close()
                except Exception:
                    pass

    # ---------- per-CUFE updates ----------

    def mark_downloaded(
        self,
        cufe: str,
        ok: bool = True,
        error: Optional[str] = None,
    ) -> None:
        """Marca una fila como Descargado / Error en memoria.

        Thread-safe. Timestamp y mensaje de error van a app.log, no al Excel.
        """
        if self._closed:
            return

        with self._lock:
            row_idx = self._cufe_to_row.get(cufe)
            if row_idx is None:
                return

            now_iso = datetime.now().isoformat(timespec="seconds")
            if ok:
                self._ws.cell(row=row_idx, column=self._col_estado).value = STATUS_DESCARGADO
                log.info("downloaded cufe=%s at=%s", cufe[:16], now_iso)
            else:
                self._ws.cell(row=row_idx, column=self._col_estado).value = STATUS_ERROR
                log.warning(
                    "error cufe=%s at=%s error=%s", cufe[:16], now_iso, error or "unknown"
                )

            self._dirty_count += 1
            dirty = self._dirty_count

        if dirty >= _FLUSH_EVERY_N_CHANGES:
            self._save_threadsafe()

    # ---------- internals ----------

    def _ensure_column(self, headers: list[str], name: str) -> int:
        """Retorna 1-based col index. Si no existe, la crea al final."""
        for i, h in enumerate(headers):
            if h == name:
                return i + 1
        new_col = len(headers) + 1
        self._ws.cell(row=1, column=new_col).value = name
        headers.append(name)
        return new_col

    def _schedule_timer(self) -> None:
        if self._closed:
            return
        self._timer = threading.Timer(_FLUSH_INTERVAL_S, self._on_timer)
        self._timer.daemon = True
        self._timer.start()

    def _on_timer(self) -> None:
        self._save_threadsafe()
        self._schedule_timer()

    def _save_threadsafe(self) -> None:
        with self._lock:
            if self._closed or self._dirty_count == 0:
                return
            try:
                self._save()
            except Exception:
                pass

    def _save(self) -> None:
        """Save atómico: save a .tmp + replace."""
        if not self._wb:
            return
        tmp = self.path + ".tmp"
        self._wb.save(tmp)
        os.replace(tmp, self.path)
        self._dirty_count = 0

    # ---------- stats ----------

    def get_pending_cufes(self) -> list[str]:
        """Retorna los CUFEs cuyo Estado actual es Pendiente o Error."""
        if not self._ws:
            return []
        pending: list[str] = []
        with self._lock:
            for cufe, row_idx in self._cufe_to_row.items():
                val = self._ws.cell(row=row_idx, column=self._col_estado).value
                estado = (val or "").strip() if val else STATUS_PENDIENTE
                if estado in (STATUS_PENDIENTE, STATUS_ERROR):
                    pending.append(cufe)
        return pending

    def status_counts(self) -> dict[str, int]:
        """Retorna {Pendiente, Descargado, Error} — útil para resumen final."""
        counts = {STATUS_PENDIENTE: 0, STATUS_DESCARGADO: 0, STATUS_ERROR: 0}
        if not self._ws:
            return counts
        with self._lock:
            for row_idx in self._cufe_to_row.values():
                val = self._ws.cell(row=row_idx, column=self._col_estado).value
                val = (val or "").strip() if val else STATUS_PENDIENTE
                if val in counts:
                    counts[val] += 1
        return counts

    @property
    def total_rows(self) -> int:
        return len(self._cufe_to_row)
